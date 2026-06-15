import openpyxl

def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    sheets = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        sheets[sheet] = data

    return {"sheets": sheets}